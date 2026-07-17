from __future__ import annotations

import io
from dataclasses import asdict

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import GeneralContact, SchoolSearchResult, StaffContact

EXPORT_COLUMNS = [
    "Input School Name", "School", "Canonical School Name", "District", "Municipality", "County",
    "Staff Name", "Exact Role", "Professional Category", "Credentials or Licenses",
    "Email", "Contact Type", "Source URL",
    "Page Title", "Evidence Snippet", "Email Confidence", "School Match Confidence",
    "Website Confidence", "Official Website", "Website Source", "Pages Searched", "Status", "Date Checked",
]


def _base_row(result: SchoolSearchResult) -> dict[str, object]:
    school = result.school
    website = result.website_match
    return {
        "Input School Name": result.input_school_name,
        "School": school.canonical_name if school else result.input_school_name,
        "Canonical School Name": school.canonical_name if school else "",
        "District": school.district_name if school else "",
        "Municipality": school.municipality if school else "",
        "County": school.county if school else "",
        "School Match Confidence": result.school_match.confidence,
        "Website Confidence": website.confidence if website else "",
        "Official Website": website.url if website else "",
        "Website Source": website.source if website else "",
        "Pages Searched": result.pages_searched,
        "Status": result.status,
        "Date Checked": result.checked_at,
    }


def _contact_row(result: SchoolSearchResult, contact: StaffContact | GeneralContact) -> dict[str, object]:
    row = _base_row(result)
    if isinstance(contact, StaffContact):
        row.update({
            "Staff Name": contact.staff_name,
            "School": contact.school or row["School"],
            "District": contact.district or row["District"],
            "Exact Role": contact.role,
            "Professional Category": contact.role_category,
            "Credentials or Licenses": contact.credentials,
            "Email": contact.email,
            "Contact Type": contact.contact_type,
            "Source URL": " | ".join(contact.source_urls),
            "Page Title": contact.page_title,
            "Evidence Snippet": contact.evidence_snippet,
            "Email Confidence": contact.email_confidence,
        })
    else:
        row.update({
            "Staff Name": "",
            "Exact Role": contact.department_name,
            "Professional Category": contact.role_category,
            "Credentials or Licenses": "",
            "Email": contact.email,
            "Contact Type": contact.contact_type,
            "Source URL": " | ".join(contact.source_urls),
            "Page Title": contact.page_title,
            "Evidence Snippet": contact.evidence_snippet,
            "Email Confidence": contact.email_confidence,
        })
    return {column: row.get(column, "") for column in EXPORT_COLUMNS}


def contacts_dataframe(results: list[SchoolSearchResult], verified_only: bool = False) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for result in results:
        contacts = result.contacts if verified_only else [*result.contacts, *result.general_contacts, *result.review_contacts]
        rows.extend(_contact_row(result, contact) for contact in contacts)
    return pd.DataFrame(rows, columns=EXPORT_COLUMNS)


def unresolved_dataframe(results: list[SchoolSearchResult]) -> pd.DataFrame:
    rows = []
    for result in results:
        if result.status in {"School match needs review", "Website needs confirmation", "Website not found", "Failed with recoverable error"}:
            row = _base_row(result)
            row["Reason"] = (result.website_match.reason if result.website_match else result.school_match.reason)
            row["Potential Candidates"] = " | ".join(candidate.url for candidate in (result.website_match.candidates if result.website_match else []))
            rows.append(row)
    return pd.DataFrame(rows)


def csv_bytes(frame: pd.DataFrame) -> bytes:
    return frame.to_csv(index=False).encode("utf-8-sig")


def excel_bytes(results: list[SchoolSearchResult]) -> bytes:
    output = io.BytesIO()
    verified_rows = []
    general_rows = []
    review_rows = []
    summary_rows = []
    for result in results:
        verified_rows.extend(_contact_row(result, contact) for contact in result.contacts)
        general_rows.extend(_contact_row(result, contact) for contact in result.general_contacts)
        review_rows.extend(_contact_row(result, contact) for contact in result.review_contacts)
        summary = _base_row(result)
        summary.update({
            "Relevant Pages": result.relevant_pages,
            "Pages Discovered": result.total_pages_discovered,
            "Page Limit Reached": result.page_limit_reached,
            "PDFs Inspected": result.pdfs_inspected,
            "Contacts Found": len(result.contacts) + len(result.general_contacts),
            "JavaScript Used": result.javascript_used,
            "Crawl Restricted": result.crawl_restricted,
        })
        summary_rows.append(summary)
    sheets = {
        "Verified Contacts": pd.DataFrame(verified_rows, columns=EXPORT_COLUMNS),
        "General Contacts": pd.DataFrame(general_rows, columns=EXPORT_COLUMNS),
        "Needs Review": pd.DataFrame(review_rows, columns=EXPORT_COLUMNS),
        "Unresolved Schools": unresolved_dataframe(results),
        "Search Summary": pd.DataFrame(summary_rows),
    }
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name, index=False)
            sheet = writer.sheets[name]
            sheet.freeze_panes = "A2"
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
            for index, column in enumerate(frame.columns, start=1):
                values = [str(column), *(str(value) for value in frame[column].head(200).tolist())]
                width = min(55, max(12, max(len(value) for value in values) + 2))
                sheet.column_dimensions[get_column_letter(index)].width = width
            sheet.auto_filter.ref = sheet.dimensions
    return output.getvalue()
